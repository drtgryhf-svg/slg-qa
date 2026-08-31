#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
lqa_scan.py — SLG 国际服本地化文本自动质检扫描器

把 LQA 清单里"机器可查"的项目一次性扫完（对应 references/lqa-checklist.md A 节）：
  placeholder   占位符/标签完整性（{0} {name} {{var}} %s %1$s <tag>）
  untranslated  未翻译串（空 / 与源文相同 / 仅大小写不同）
  glossary      术语一致性（源文含 src_term 时，译文必须用 tgt_term，禁用 forbidden）
  length        文本超长（按语言膨胀系数，或 maxlen 列/参数硬上限）
  whitespace    首尾空格、连续空格
  punct        重复标点（！！ 。。 等，"..." 除外）
  escape        \\n \\t 引号转义数量与源文不一致
  encoding      U+FFFD 乱码、bidi 控制符、全/半角标点混用
  casing        全大写风格与源文不一致（仅拉丁语言）

用法（详见 SKILL.md）:
  python lqa_scan.py strings.csv --src-col en --tgt-col de --lang de --glossary glossary.csv
  python lqa_scan.py strings.xlsx --sheet Sheet1 --src-col A --tgt-col C --lang ja
  python lqa_scan.py i18n.json --src en --tgt de --lang de          # {"key": {"en":..., "de":...}}
  python lqa_scan.py strings_en.xml strings_de.xml --src-file strings_en.xml --tgt-file strings_de.xml --lang de

输入格式: CSV / XLSX / JSON(dict 或 list) / Android strings.xml
输出: text（默认）/ md / json；退出码 0=无发现 1=有发现

严重等级提示是参考默认值（S2/S3/S4），最终等级由 QA 依据 bug-and-report-format.md 裁定。
"""

import argparse
import csv
import io
import json
import re
import sys
from collections import Counter
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ---------------------------------------------------------------- 常量

PLACEHOLDER_RE = re.compile(
    r"\{\{[^{}]+\}\}"          # {{var}}
    r"|\{[^{}]+\}"             # {0} {name}
    r"|%(?:\d+\$)?[sdf]"       # %s %d %f %1$s
    r"|</?[a-zA-Z][^<>]*>"     # <b> </color=xxx> 等富文本标签
)
TAG_NAME_RE = re.compile(r"</?([a-zA-Z][a-zA-Z0-9_-]*)")
BIDI_RE = re.compile(r"[\u200e\u200f\u202a-\u202e\u2066-\u2069]")
CJK_RE = re.compile(r"[\u3400-\u9fff\uf900-\ufaff\u3040-\u30ff\uac00-\ud7af]")

# 目标语言文本膨胀系数上限（相对源文长度）。带 --max-ratio 或 --maxlen 覆盖。
LENGTH_LIMITS = {
    "de": 1.40, "nl": 1.40, "fi": 1.30, "hu": 1.30, "ar": 1.30, "vi": 1.30,
    "fr": 1.25, "es": 1.25, "pt": 1.25, "it": 1.25, "sv": 1.25, "da": 1.25,
    "no": 1.25, "nb": 1.25, "pl": 1.25, "id": 1.25, "ms": 1.25, "he": 1.25,
    "ru": 1.20, "uk": 1.20, "tr": 1.20, "cs": 1.20, "el": 1.20, "th": 1.20,
    "ro": 1.25, "en": 1.00, "ja": 1.00, "ko": 1.00, "zh": 1.00,
}

SEVERITY = {
    "placeholder": "S2", "untranslated": "S2", "glossary": "S3",
    "length": "S3", "whitespace": "S4", "punct": "S4",
    "escape": "S4", "encoding": "S4", "casing": "S4",
}
CHECK_NAMES = {
    "placeholder": "占位符/标签", "untranslated": "未翻译", "glossary": "术语",
    "length": "超长", "whitespace": "空格", "punct": "重复标点",
    "escape": "转义", "encoding": "编码", "casing": "大小写风格",
}
FW_PUNCT = set("！？：；，（）。")


# ---------------------------------------------------------------- 读取

def read_text(path):
    raw = Path(path).read_bytes()
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return raw.decode("utf-16")
    for enc in ("utf-8-sig", "utf-8", "gbk", "cp1252"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def sniff_delim(sample):
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except Exception:
        return ","


def resolve_col(spec, header, default=0, allow_letters=False):
    """列参数: 表头名 / 1 起始序号；仅 XLSX 下表头匹配失败时再按 Excel 字母(A, BC)解析，
    避免吞掉 en/de/id 这类短列名。"""
    if spec is None:
        return default
    spec = str(spec).strip()
    if re.fullmatch(r"\d+", spec):
        return int(spec) - 1
    for i, h in enumerate(header):
        if str(h).strip() == spec:
            return i
    for i, h in enumerate(header):
        if str(h).strip().casefold() == spec.casefold():
            return i
    if allow_letters and re.fullmatch(r"[A-Za-z]{1,3}", spec):
        n = 0
        for ch in spec.upper():
            n = n * 26 + (ord(ch) - 64)
        return n - 1
    raise SystemExit(f"错误: 找不到列 '{spec}'，表头为: {header}")


def cell(v):
    return "" if v is None else str(v)


def load_csv(path, args):
    text = read_text(path)
    delim = sniff_delim(text[:4096])
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    if not rows:
        return []
    header = rows[0]
    si = resolve_col(args.src_col, header)
    ti = resolve_col(args.tgt_col, header)
    ki = resolve_col(args.key_col, header, default=None) if args.key_col else _guess_key_col(header)
    mi = resolve_col(args.maxlen_col, header, default=None) if args.maxlen_col else None
    out = []
    for i, r in enumerate(rows[1:], start=2):
        if not any(c.strip() for c in r):
            continue
        out.append({
            "key": cell(r[ki]) if ki is not None and ki < len(r) else f"row{i}",
            "src": cell(r[si]) if si < len(r) else "",
            "tgt": cell(r[ti]) if ti < len(r) else "",
            "maxlen": cell(r[mi]).strip() if mi is not None and mi < len(r) else "",
            "file": str(path), "row": i,
        })
    return out


def _guess_key_col(header):
    for i, h in enumerate(header):
        if str(h).strip().casefold() in ("key", "id", "string_id", "stringid", "name"):
            return i
    return 0


def load_xlsx(path, args):
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("错误: 读 XLSX 需要 openpyxl（pip install openpyxl），或改用 CSV。")
    from openpyxl.utils import column_index_from_string
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if args.sheet:
        ws = wb[wb.sheetnames[int(args.sheet) - 1]] if str(args.sheet).isdigit() else wb[args.sheet]
    else:
        ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [cell(c).strip() for c in next(it)]
    si = resolve_col(args.src_col, header, allow_letters=True)
    ti = resolve_col(args.tgt_col, header, allow_letters=True)
    ki = resolve_col(args.key_col, header, default=None, allow_letters=True) if args.key_col else _guess_key_col(header)
    mi = resolve_col(args.maxlen_col, header, default=None, allow_letters=True) if args.maxlen_col else None
    out = []
    for i, r in enumerate(it, start=2):
        vals = [cell(c) for c in r]
        if not any(v.strip() for v in vals):
            continue
        out.append({
            "key": vals[ki] if ki is not None and ki < len(vals) else f"row{i}",
            "src": vals[si] if si < len(vals) else "",
            "tgt": vals[ti] if ti < len(vals) else "",
            "maxlen": vals[mi].strip() if mi is not None and mi < len(vals) else "",
            "file": f"{Path(path).name}::{ws.title}", "row": i,
        })
    return out


def load_json_records(path, args):
    data = json.loads(read_text(path))
    out = []
    if isinstance(data, dict) and data and all(isinstance(v, dict) for v in data.values()):
        # {"key": {"en": "...", "de": "..."}}
        for k, v in data.items():
            out.append({
                "key": k, "src": cell(v.get(args.src, "")), "tgt": cell(v.get(args.tgt, "")),
                "maxlen": "", "file": str(path), "row": k,
            })
    elif isinstance(data, dict):
        # {"key": "text"} 单语，需要 --src-file/--tgt-file 双文件模式
        for k, v in data.items():
            out.append({"key": k, "src": cell(v), "tgt": "", "maxlen": "",
                        "file": str(path), "row": k})
    elif isinstance(data, list):
        for i, it in enumerate(data, start=1):
            if not isinstance(it, dict):
                raise SystemExit(f"错误: JSON list 元素应为对象: {path}#{i}")
            k = cell(it.get("key", it.get("id", i)))
            s = cell(it.get(args.src_col or "source", ""))
            t = cell(it.get(args.tgt_col or "target", ""))
            out.append({"key": k, "src": s, "tgt": t, "maxlen": "",
                        "file": str(path), "row": i})
    else:
        raise SystemExit(f"错误: 不支持的 JSON 结构: {path}")
    return out


def load_xml_strings(path):
    import xml.etree.ElementTree as ET
    root = ET.fromstring(read_text(path))
    return {el.get("name") or "": (el.text or "") for el in root.iter("string")}


def load_two_files(args):
    """双文件模式: --src-file X --tgt-file Y（JSON dict 或 Android strings.xml）"""
    def load_one(p, lang):
        text = read_text(p)
        if p.lower().endswith(".xml"):
            return load_xml_strings(p)
        d = json.loads(text)
        if not isinstance(d, dict):
            raise SystemExit(f"错误: 双文件模式的 JSON 须为 {{key: text}}: {p}")
        if lang and d and all(isinstance(v, dict) for v in d.values()):
            return {k: cell(v.get(lang, "")) for k, v in d.items()}
        return {k: cell(v) for k, v in d.items()}

    if not (args.src_file and args.tgt_file):
        raise SystemExit("错误: 双文件模式需要同时给 --src-file 和 --tgt-file")
    src = load_one(args.src_file, args.src)
    tgt = load_one(args.tgt_file, args.tgt)
    out = []
    for k in sorted(set(src) | set(tgt)):
        out.append({"key": k, "src": src.get(k, ""), "tgt": tgt.get(k, ""),
                    "maxlen": "",
                    "missing": ("tgt" if k not in tgt else "src" if k not in src else ""),
                    "file": f"{args.src_file} + {args.tgt_file}", "row": k})
    return out


def load_glossary(path, lang):
    """CSV: src_term,tgt_term[,lang][,forbidden(分号分隔)]；或 term,action(required|forbidden)"""
    text = read_text(path)
    delim = sniff_delim(text[:2048])
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    if not rows:
        return []
    header = [h.strip().casefold() for h in rows[0]]
    rules = []
    for r in rows[1:]:
        if not any(c.strip() for c in r):
            continue
        if "src_term" in header:
            g = dict(zip(header, [c.strip() for c in r]))
            rules.append({
                "src_term": g.get("src_term", ""), "tgt_term": g.get("tgt_term", ""),
                "lang": g.get("lang", ""), "forbidden": g.get("forbidden", ""),
            })
        elif len(r) >= 2 and r[1].strip().lower() in ("required", "forbidden"):
            rules.append({"src_term": "", "tgt_term": r[0].strip(),
                          "lang": r[2].strip() if len(r) > 2 else "",
                          "forbidden": r[0].strip() if r[1].strip().lower() == "forbidden" else ""})
    if lang:
        rules = [g for g in rules if not g["lang"] or g["lang"].lower() == lang.lower()]
    return rules


# ---------------------------------------------------------------- 检查

def contains(text, term):
    t = term.strip()
    if not t:
        return False
    if re.search(r"\w", t[0], re.UNICODE) and re.search(r"\w", t[-1], re.UNICODE):
        return re.search(r"(?<!\w)" + re.escape(t) + r"(?!\w)", text, re.IGNORECASE) is not None
    return t.lower() in text.lower()


def check_placeholder(r):
    st = Counter(re.sub(r"\s+", "", t) for t in PLACEHOLDER_RE.findall(r["src"]))
    tt = Counter(re.sub(r"\s+", "", t) for t in PLACEHOLDER_RE.findall(r["tgt"]))
    if not st and not tt:
        return []
    missing = st - tt
    extra = tt - st
    if not missing and not extra:
        # 占位符集合相同，再核对富文本标签名配对
        sn = Counter(TAG_NAME_RE.findall(r["src"]))
        tn = Counter(TAG_NAME_RE.findall(r["tgt"]))
        if sn != tn:
            diff = {k for k in set(sn) | set(tn) if sn.get(k, 0) != tn.get(k, 0)}
            return [_f(r, f"富文本标签名不一致: {sorted(diff)} (src={dict(sn)} tgt={dict(tn)})")]
        return []
    parts = []
    if missing:
        parts.append("缺少 " + " ".join(f"{t}×{n}" if n > 1 else t for t, n in sorted(missing.items())))
    if extra:
        parts.append("多出 " + " ".join(f"{t}×{n}" if n > 1 else t for t, n in sorted(extra.items())))
    return [_f(r, "；".join(parts))]


def check_untranslated(r):
    s, t = r["src"].strip(), r["tgt"].strip()
    if r.get("missing") == "tgt":
        return [_f(r, "该 key 在目标语言文件中不存在（整条漏翻）")]
    if r.get("missing") == "src":
        return [_f(r, "该 key 在源语言文件中不存在（目标多出条目，可能是废弃串残留或源文缺失）")]
    if not s:
        return []
    if not t:
        return [_f(r, "译文为空（未翻译）")]
    if re.fullmatch(r"[\d\s.,:%+/\-–—°'\"~×xX]+", s):
        return []  # 纯数字/符号串不需要翻译
    if t == s:
        return [_f(r, "目标与源文完全相同")]
    if t.casefold() == s.casefold():
        return [_f(r, "目标与源文仅大小写不同")]
    return []


def check_glossary(r, rules):
    out = []
    if not r["src"].strip() or not r["tgt"].strip():
        return out
    for g in rules:
        st, tt = g["src_term"].strip(), g["tgt_term"].strip()
        if st and not contains(r["src"], st):
            continue
        if tt and not contains(r["tgt"], tt):
            out.append(_f(r, f"术语缺失: 源文含 '{st}' 但译文未用规定译法 '{tt}'"))
        for bad in [x.strip() for x in g["forbidden"].split(";") if x.strip()]:
            if contains(r["tgt"], bad):
                out.append(_f(r, f"禁用术语: 译文出现 '{bad}'（应使用 '{tt or st}'）"))
    return out


def check_length(r, lang, max_ratio, hard_max):
    s, t = r["src"].strip(), r["tgt"].strip()
    if not s or not t or CJK_RE.search(s):
        return []
    if r.get("maxlen", "").isdigit():
        if len(t) > int(r["maxlen"]):
            return [_f(r, f"超出 max-length: {len(t)} > {r['maxlen']} 字符")]
        return []
    if hard_max and len(t) > hard_max:
        return [_f(r, f"超出 --maxlen 硬上限: {len(t)} > {hard_max} 字符")]
    lim = max_ratio or LENGTH_LIMITS.get((lang or "").lower(), 1.25)
    ratio = len(t) / max(len(s), 1)
    if ratio > lim:
        return [_f(r, f"长度比 {ratio:.2f} > 上限 {lim:.2f} (语言: {lang or '未指定，默认1.25'}), "
                       f"src {len(s)} 字符 → tgt {len(t)} 字符")]
    return []


def check_whitespace(r):
    t = r["tgt"]
    if not t.strip():
        return []
    out = []
    if t != t.strip():
        out.append(_f(r, "译文首/尾有空格"))
    if "  " in t:
        out.append(_f(r, f"连续空格 ×{t.count('  ')}（注意: 双空格也可能是刻意排版，需人工确认）"))
    return out


def check_punct(r):
    t = r["tgt"]
    if not t.strip():
        return []
    out = []
    for p in ("。。", "！！", "？？", "、、", "，，", "；；", "：：",
              "!!", "??", ",,", ";;", "::"):
        if p in t:
            out.append(_f(r, f"重复标点 '{p}'（若为刻意语气风格可人工忽略）"))
    if re.search(r"(?<!\.)\.\.(?!\.)", t):
        out.append(_f(r, "连续两个句点 '..'（省略号应为 '...' 或 '…'）"))
    return out


def check_escape(r):
    s, t = r["src"], r["tgt"]
    if not s.strip() or not t.strip():
        return []
    out = []
    for name, pat in (("\\n", "\\n"), ("\\t", "\\t"), ('\\"', '\\"')):
        if s.count(pat) != t.count(pat):
            out.append(_f(r, f"转义符 {name} 数量不一致: src×{s.count(pat)} tgt×{t.count(pat)}"))
    if s.count('"') % 2 != t.count('"') % 2:
        out.append(_f(r, f"引号数量奇偶不一致: src×{s.count(chr(34))} tgt×{t.count(chr(34))}"))
    return out


def check_encoding(r, lang):
    t = r["tgt"]
    if not t.strip():
        return []
    out = []
    if "\ufffd" in t:
        out.append(_f(r, "存在 U+FFFD 乱码字符"))
    if BIDI_RE.search(t) and not BIDI_RE.search(r["src"]):
        out.append(_f(r, "夹带 bidi 控制符（源文没有）"))
    if lang in ("ja", "zh") and re.search(
            r"[\u3040-\u30ff\u3400-\u9fff\uac00-\ud7af][!?;:(),]|[!?;:(),][\u3040-\u30ff\u3400-\u9fff\uac00-\ud7af]", t):
        out.append(_f(r, f"半角标点紧贴 CJK 字符（{lang} 应使用全角）"))
    if lang not in ("ja", "zh", "ko", None, "") and any(c in FW_PUNCT for c in t):
        fw = "".join(c for c in t if c in FW_PUNCT)
        out.append(_f(r, f"出现全角标点 '{fw}'（拉丁语言应为半角）"))
    return out


def check_casing(r, lang):
    s, t = r["src"], r["tgt"]
    if not s.strip() or not t.strip():
        return []
    if (lang or "").lower() in ("ja", "zh", "ko", "ar", "he", "th"):
        return []
    ls = "".join(c for c in s if c.isalpha())
    lt = "".join(c for c in t if c.isalpha())
    if len(ls) < 3 or len(lt) < 3 or sum(c.isascii() for c in lt) / len(lt) < 0.6:
        return []
    if ls.isupper() and not lt.isupper():
        return [_f(r, "源文为全大写风格，译文不是")]
    if not ls.isupper() and lt.isupper():
        return [_f(r, "译文为全大写风格，源文不是")]
    return []


def _f(r, detail):
    return {"check": None, "file": r["file"], "row": r["row"], "key": r["key"],
            "detail": detail, "src": r["src"], "tgt": r["tgt"]}


CHECKS = {
    "placeholder": lambda r, a: check_placeholder(r),
    "untranslated": lambda r, a: check_untranslated(r),
    "glossary": lambda r, a: check_glossary(r, a._rules) if a._rules else [],
    "length": lambda r, a: check_length(r, a.lang, a.max_ratio, a.maxlen),
    "whitespace": lambda r, a: check_whitespace(r),
    "punct": lambda r, a: check_punct(r),
    "escape": lambda r, a: check_escape(r),
    "encoding": lambda r, a: check_encoding(r, a.lang),
    "casing": lambda r, a: check_casing(r, a.lang),
}


# ---------------------------------------------------------------- 输出

def excerpt(s, n=70):
    s = s.replace("\n", "\\n").replace("\t", "\\t")
    return s if len(s) <= n else s[:n] + "…"


def render_text(findings, args, total_records):
    lines = []
    by_check = group(findings)
    order = [c for c in CHECKS if c in by_check]
    if not order:
        lines.append(f"✓ 扫描完成: {total_records} 条文本，未发现机器可查问题。")
        return "\n".join(lines)
    lines.append(f"扫描完成: {total_records} 条文本，发现 {len(findings)} 条问题，按检查项分组:\n")
    for c in order:
        items = by_check[c]
        shown = items if args.all else items[:args.max_show]
        lines.append(f"── [{CHECK_NAMES[c]}] {c}  共 {len(items)} 条  参考等级 {SEVERITY[c]} {'─' * 20}")
        for it in shown:
            lines.append(f"  {it['file']}  行 {it['row']}  key={it['key']}")
            lines.append(f"    问题: {it['detail']}")
            if it["src"]:
                lines.append(f"    SRC: {excerpt(it['src'])}")
            lines.append(f"    TGT: {excerpt(it['tgt']) or '(空)'}")
        if len(items) > len(shown):
            lines.append(f"  …… 其余 {len(items) - len(shown)} 条用 --all 或 --max-show 查看")
        lines.append("")
    lines.append("提示: 等级为参考默认值，按 references/bug-and-report-format.md 裁定；")
    lines.append("      untranslated/glossary 条目需人工排除专有名词误报后再定级。")
    return "\n".join(lines)


def render_md(findings, args, total_records):
    out = ["# LQA 自动扫描报告\n",
           f"- 范围: {', '.join(args.files)}　语言: {args.lang or '未指定'}",
           f"- 文本条数: {total_records}　发现: {len(findings)} 条\n"]
    by_check = group(findings)
    for c in [c for c in CHECKS if c in by_check]:
        out.append(f"## {CHECK_NAMES[c]}（{c}，参考等级 {SEVERITY[c]}）— {len(by_check[c])} 条\n")
        out.append("| 文件 | 行 | key | 问题 | SRC | TGT |")
        out.append("|---|---|---|---|---|---|")
        for it in by_check[c]:
            out.append(f"| {it['file']} | {it['row']} | {it['key']} | {it['detail']} "
                       f"| {excerpt(it['src'], 40)} | {excerpt(it['tgt'], 40) or '(空)'} |")
        out.append("")
    return "\n".join(out)


def render_json(findings, args, total_records):
    return json.dumps({
        "files": args.files, "lang": args.lang, "total_records": total_records,
        "total_findings": len(findings),
        "summary": dict(Counter(f["check"] for f in findings)),
        "findings": findings,
    }, ensure_ascii=False, indent=2)


def group(findings):
    g = {}
    for f in findings:
        g.setdefault(f["check"], []).append(f)
    return g


# ---------------------------------------------------------------- 主流程

def main():
    p = argparse.ArgumentParser(description="SLG 国际服 LQA 自动扫描器（见 SKILL.md）")
    p.add_argument("files", nargs="*", help="文本表文件 (CSV/XLSX/JSON/XML)")
    p.add_argument("--src-col", help="源文列: 表头名/序号/Excel字母（JSON list 模式为字段名）")
    p.add_argument("--tgt-col", help="目标文列: 同上")
    p.add_argument("--key-col", help="key 列（默认自动识别 key/id/name，否则第 1 列）")
    p.add_argument("--sheet", help="XLSX 工作表名或 1 起始序号")
    p.add_argument("--src", help="JSON dict 模式的源语言键名 / 双文件模式源语言代码")
    p.add_argument("--tgt", help="JSON dict 模式的目标语言键名 / 双文件模式目标语言代码")
    p.add_argument("--src-file", help="双文件模式: 源语言文件（JSON dict / strings.xml）")
    p.add_argument("--tgt-file", help="双文件模式: 目标语言文件")
    p.add_argument("--lang", help="目标语言代码 (de/ja/...)，用于长度系数与术语表")
    p.add_argument("--glossary", help="术语表 CSV")
    p.add_argument("--max-ratio", type=float, help="覆盖长度比上限")
    p.add_argument("--maxlen", type=int, help="译文硬性字符上限")
    p.add_argument("--maxlen-col", help="CSV/XLSX 中逐行 max-length 列名")
    p.add_argument("--checks", default="all",
                   help="只跑部分检查，逗号分隔: placeholder,untranslated,glossary,length,whitespace,punct,escape,encoding,casing")
    p.add_argument("--format", choices=["text", "md", "json"], default="text")
    p.add_argument("--out", help="报告输出文件路径")
    p.add_argument("--all", action="store_true", help="显示全部发现（默认每组最多 --max-show 条）")
    p.add_argument("--max-show", type=int, default=50)
    args = p.parse_args()

    args._rules = load_glossary(args.glossary, args.lang) if args.glossary else []

    # 组装记录
    records = []
    if args.src_file or args.tgt_file:
        records = load_two_files(args)
    elif not args.files:
        p.error("请提供输入文件，或用 --src-file/--tgt-file 双文件模式")
    for f in args.files:
        fl = f.lower()
        if fl.endswith(".csv"):
            records += load_csv(f, args)
        elif fl.endswith((".xlsx", ".xlsm")):
            records += load_xlsx(f, args)
        elif fl.endswith(".json"):
            if not (args.src and args.tgt):
                raise SystemExit("错误: JSON 文件需 --src 与 --tgt 指定语言键名")
            records += load_json_records(f, args)
        elif fl.endswith(".xml"):
            raise SystemExit("错误: XML 请用双文件模式 --src-file/--tgt-file")
        else:
            raise SystemExit(f"错误: 不支持的文件类型: {f}（支持 CSV/XLSX/JSON/XML）")

    # 执行检查
    wanted = list(CHECKS) if args.checks == "all" else [c.strip() for c in args.checks.split(",")]
    bad = [c for c in wanted if c not in CHECKS]
    if bad:
        p.error(f"未知检查项: {bad}（可选: {', '.join(CHECKS)}）")

    findings = []
    for r in records:
        for c in wanted:
            for x in CHECKS[c](r, args):
                x["check"] = c
                findings.append(x)

    total = len(records)
    text = {"text": render_text, "md": render_md, "json": render_json}[args.format](findings, args, total)
    if args.out:
        Path(args.out).write_text(text + "\n", encoding="utf-8")
        print(f"报告已写入 {args.out}")
    print(text)
    sys.exit(1 if findings else 0)


if __name__ == "__main__":
    main()
